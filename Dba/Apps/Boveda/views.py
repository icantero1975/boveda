from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.template.loader import get_template
from Dba.Apps.Boveda.form import MpForm, ExpedienteForm, IndiciosForm, DefForm,TemForm,TrasForm
from Dba.Apps.Boveda.models import Mp,Expediente,Indicio,Definitiva,Temporal,Traslado
from django.views.generic import ListView, CreateView,UpdateView,DeleteView
from django.template import Context
from django.http import Http404
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image
from openpyxl.workbook import Workbook






# Create your views here.
from django.contrib.auth import logout

def logout_view(request):
    logout(request)
    return render(request,'boveda/login.html')
    # Redirect to a success page.

def index(request):
    return render(request,'boveda/index.html')

def mp_view(request):
    if request.method == 'POST':
        form = MpForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('autoridad_listar')
    else:
        form = MpForm()
    return render(request,'boveda/mp_form.html',{'form':form})

def mp_list(request):
    autoridad = Mp.objects.all().order_by('id')
    contexto = {'autoridades': autoridad}
    return render(request, 'boveda/Mp_list.html',contexto)

def mp_edit(request,id_mp):
    autoridad = Mp.objects.get(id=id_mp)
    if request.method == 'GET':
        form = MpForm(instance=autoridad)
    else:
        form = MpForm(request.POST, instance=autoridad)
        if form.is_valid():
            form.save()
        return redirect('autoridad_listar')
    return render(request,'boveda/mp_form.html',{'form':form})

def mp_delete(request,id_mp):
    autoridad = Mp.objects.get(id=id_mp)
    if request.method == 'POST':
        autoridad.delete()
        return redirect('autoridad_listar')
    return render(request,'boveda/mp_delete.html',{'autoridad':autoridad})

class Autoridadlist(ListView):
    model = Mp
    template_name = 'Mp_list.html'
    paginate_by = 8

class AutoridadCreate(CreateView):
    model = Mp
    form_class = MpForm
    template_name = 'boveda/mp_form.html'
    success_url = reverse_lazy('autoridad_listar')

class AutoridadUpdateView(UpdateView):
    model = Mp
    form_class = MpForm
    template_name = 'boveda/mp_form.html'
    success_url = reverse_lazy('autoridad_listar')

class AutoridadDeleteView(DeleteView):
    model = Mp
    template_name = 'boveda/mp_delete.html'
    success_url = reverse_lazy('autoridad_listar')

class ExpedienteList(ListView):
    model = Expediente
    template_name = 'expediente_list.html'
    paginate_by = 5

class ExpedienteCreate(CreateView):
    model = Expediente
    form_class = ExpedienteForm
    template_name = 'boveda/expediente_form.html'
    success_url = reverse_lazy('expediente_listar')

class ExpedienteUpdateView(UpdateView):
    model = Expediente
    form_class = ExpedienteForm
    template_name = 'boveda/expediente_form.html'
    success_url = reverse_lazy('expediente_listar')

class IndicioList(ListView):
    model = Indicio
    template_name = 'boveda\indicios_list.html'
    paginate_by = 10

class IndicioCreate(CreateView):
    model = Indicio
    form_class = IndiciosForm
    template_name = 'boveda/indicios_form.html'
    success_url = reverse_lazy('indicio_listar')

class IndicioUpdateView(UpdateView):
    model = Indicio
    form_class = IndiciosForm
    template_name = 'boveda/indicios_form.html'
    success_url = reverse_lazy('indicio_listar')

# crear las funciones de vista para el modelo de salida definitiva

class DefinitivaCreate(CreateView):
    model = Definitiva
    form_class = DefForm
    template_name = 'boveda/def_form.html'
    success_url = reverse_lazy('def_listar')

class DefinitivaList(ListView):
    model = Definitiva
    template_name = 'boveda/def_list.html'
    paginate_by = 10

class DefinitivaUpdateView(UpdateView):
    model = Definitiva
    form_class = DefForm
    template_name = 'boveda/def_form.html'
    success_url = reverse_lazy('def_listar')

# crear las clases de vista para el modelo de salidas temporales

class TemporalCreate(CreateView):
    model = Temporal
    form_class = TemForm
    template_name = 'boveda/temp_form.html'
    success_url = reverse_lazy('tem_listar')

class TemporalList(ListView):
    model = Temporal
    template_name = 'boveda/temp_list.html'
    paginate_by = 10

class TemporalUpdateView(UpdateView):
    model = Temporal
    form_class = TrasForm
    template_name = 'boveda/temp_form.html'
    success_url = reverse_lazy('tem_listar')

# crear las clases de vista para el modelo traslados a la bodega 
class TrasladoCreate(CreateView):
    model = Traslado
    form_class = TrasForm
    template_name = 'boveda/tras_form.html'
    success_url = reverse_lazy('tras_listar')

class TrasladoList(ListView):
    model = Traslado
    template_name = 'boveda/tras_list.html'
    paginate_by = 10

class TrasladoUpdateView(UpdateView):
    model = Traslado
    form_class = TrasForm
    template_name = 'boveda/tras_form.html'
    success_url = reverse_lazy('tras_listar')


# crear la funcion para buscar los indicios de una entrada o carpeta y ponerlos en un diccionario
# renderizarlo y mandarlo a un PDF para que se pueda imprimir

def buscaindicios(request,id_carpeta):

    try:
        indicio = Indicio.objects.get(pk=id_carpeta)
    except indicio.DoesNotExist:
        raise Http404("El indicio no existe")

        carpeta = Expediente.objects.filter(indicio=indicio)

        # Datos para enviar al render del template (via context)
    data = {'indicio': indicio,
            'expediente': carpeta,
            }
    return data

def mis_indicios(request,q):
    #q = request.GET.get('q', '')

    indicios = Indicio.objects.filter(Ingreso_Num=q)
    return render(request,'boveda/formato_list.html', {'indicios': indicios})

class ReporteTraslado(TemplateView):

        #Usamos el método get para generar el archivo excel
        def get(self, request, *args, **kwargs):
                #Obtenemos todas las personas de nuestra base de datos

                campo = int(request.GET.get('campo'))
                indicios = Traslado.objects.filter(Carpeta=campo)
                #Creamos el libro de trabajo
                wb = Workbook()
                #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
                ws = wb.active
                # pongo la imagen
                img = Image('media/imagen/fiscalia.png')
                ws.add_image(img, 'A1')
                # Hacer encabezado
                #le pongo el nombre de la hoja con la carpeta
                #ws.title = indicios.Ingreso_Num.Carpeta_Investigacion
                # para obtener el numero de carpeta, hora, fecha y quien entrega por una vez

                
                for medio in indicios:
                        fecha= medio.Fecha_de_Movimiento
                        carpeta = medio.Carpeta.Carpeta_Investigacion
                        autoridad = medio.Autoridad.Nombre
                        persona =  medio.Persona_Recibe
                        

                # nombre de la hoja como la carpeta
                #date_object = datetime.strptime(fecha, '%dd/%mm/%Y')
                fecha = datetime.strftime(fecha, '%d/%m/%Y')       
                ws.title = 'Traslado'
                #En la celda C2 ponemos el texto 'FISCALIA GENERAL DEL ESTADO' y la decoramos
                #alignment=Alignment
                ws['C3'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C3'].font = Font(name='Arial',size = 14, bold= True)
                ws['C3'] = 'Fiscalía General del Estado'
                #En la celda C3 ponemos el texto 'Oficialía Mayor' y la decoramos
                ws['C4'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C4'].font = Font(name='Arial',size = 11, bold= True, color = "BFBDA4")
                ws['C4'] = 'Oficialía Mayor'
                # en la celda c4 ponemos el texto "Dirección de Bienes Asegurados"
                ws['C5'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C5'].font = Font(name='Arial',size = 11, bold= True)
                ws['C5'] = 'Dirección de Bienes Asegurados'
                # en la celda C6 ponemos el Texto "Traslado de Indicios a la Bodega de Bienes Asegurados"
                ws.merge_cells('A6:F6')
                ws['A6'].alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws['A6'].font = Font(name='Arial',size = 12, bold= True)
                ws['A6'] = 'TRASLADO DE INDICIOS A LA BODEGA DE BIENES ASEGURADOS'
                # En la celda A8:B8 pones el titulo "Carpeta de Investigacion" con relleno y bordes
                ws.merge_cells('A8:B8')
                ws['A8'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['A8'].font = Font(name='Arial',size = 10, bold= True,color ="EAEAEA")
                ws['A8'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['A8'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['A8'] = 'CARPETA DE INVESTIGACION'
                # En la celda COMBINADA pones la variable carpeta sin  relleno y bordes
                ws.merge_cells('A9:B11')
                ws['A9'].alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws['A9'].font = Font(name='Arial',size = 10, bold= False)
                ws['A9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['A10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['A11'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['B11'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=9,column=1).value = carpeta
                # En la celda C8 ponemos el titulo Fecha 

                ws['C8'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C8'].font = Font(name='Arial',size = 10, bold= True,color ="EAEAEA")
                ws['C8'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['C8'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['C8'] = 'FECHA'
                # EN LA celda C9 ponemos el valor de la Fecha

                ws.merge_cells('C9:C11')
                ws['C9'].alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws['C9'].font = Font(name='Arial',size = 10, bold= False)
                ws['C9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['C10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['C11'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=9,column=3).value = fecha 
                # en la celda D8 ponemos el titulo de la Hora de Entrega

                # COMBINAMOS LAS CELDAS PARA PONER EL ENCABEZADO "MARQUE CON UNA X" 
                ws.merge_cells('D8:E8')
                ws['D8'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['D8'].font = Font(name='Arial',size = 10, bold= True,color ="EAEAEA")
                ws['D8'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['D8'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['D8'] = 'MARQUE CON X'
                # PONEMOS EL TEXTO COPIA BOVEDA DE INDICIOS EN LA CELDA E9,
                ws['D9'].alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws['D9'].font = Font(name='Arial',size = 8, bold= False)
                ws['D9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['D9'] = 'COPIA BOVEDA DE INDICIOS'
                # PONEMOS EL TEXTO COPIA BOVEDA DE INDICIOS EN LA CELDA E10,
                ws['D10'].alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws['D10'].font = Font(name='Arial',size = 8, bold= False)
                ws['D10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['D10'] = 'COPIA VIGILANTE SANCTORUM'
                 # PONEMOS EL TEXTO COPIA BOVEDA DE INDICIOS EN LA CELDA E11,
                ws['D11'].alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws['D11'].font = Font(name='Arial',size = 8, bold= False)
                ws['D11'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['D11'] = 'COPIA BODEGA DE BIENES ASEGURADOS'
                # PONEMOS LOS BORDES A LAS CELDAS F9,F10,F11
                ws['E9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['E10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['E11'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                # HACER EL ENCABEZADO
                #RUE
                ws['A14'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['A14'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['A14'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['A14'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['A14'] = 'RUE'
                #TIPO
                ws['B14'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['B14'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['B14'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['B14'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['B14'] = 'TIPO'
                #DESCRIPCION
                
                #ws.merge_cells('C14:D15')
                ws['C14'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C14'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['C14'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['C14'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['C14'] = 'DESCRIPCION'
                # definir la anchura para las columnas a,b,c,d,e,y f
                ws.column_dimensions['A'].width = 17
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 70
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 5
                
                # definir la altura para las filas
                ws.row_dimensions[6].height = 30
                ws.row_dimensions[9].height = 24
                ws.row_dimensions[10].height = 24
                ws.row_dimensions[11].height = 35

                cont=15
                #Recorremos el conjunto de indicios y vamos escribiendo cada uno de los datos en las celdas

                for indicio in indicios:
                        ws.cell(row=cont, column=1).value = indicio.Rue.Rue
                        ws.cell(row=cont, column=2).value = indicio.Rue.Tipo
                        ws.cell(row=cont, column=3).value = indicio.Rue.Descripcion
                        #ponemos el diseño de la celda, el tipo de letra  y los borde
                        ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                        ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                        ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                        # columna dos
                        ws.cell(row=cont, column=2).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                        ws.cell(row=cont, column=2).font = Font(name='Arial',size = 10, bold= False)
                        ws.cell(row=cont, column=2).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                        # columna 3
                        #ws.merge_cells(start_row=cont,start_column = 3, end_row = cont, end_column = 4)
                        ws.cell(row=cont, column=3).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                        ws.cell(row=cont, column=3).font = Font(name='Arial',size = 10, bold= False)
                        ws.cell(row=cont, column=3).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                        ws.row_dimensions[cont].height = 60
                        #ws.merge_cells(start_row=cont,start_column = 3, end_row = cont, end_column = 4)
                        cont += 1

                # dibujamos el marco completo

                # ponemos las firmas 
                # ENCABEZADO
                cont +=1
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= True,color ="EAEAEA")
                ws.cell(row=cont, column=1).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=1).value = "VALIDA TRASLADO DE BIENES A BODEGA DE BIENES ASEGURADOS"
                ws.row_dimensions[cont].height = 25
                #PONEMOS EL TEXTO NOMBRE 
                cont +=1
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"))
                ws.cell(row=cont, column=1).value = "NOMBRE:"
                ws.merge_cells(start_row=cont,start_column = 2, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"))
                ws.row_dimensions[cont].height = 25
                # PONEMOS EL TEXTO DE CARGO
                cont +=1
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"))
                ws.cell(row=cont, column=1).value = "CARGO:"
                ws.merge_cells(start_row=cont,start_column = 2, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"))    
                ws.row_dimensions[cont].height = 25
                #PONEMOS EL TEXTO DE FIRMA
                cont +=1
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"),bottom = Side(border_style='thin'))
                ws.cell(row=cont, column=1).value = "FIRMA:"
                ws.merge_cells(start_row=cont,start_column = 2, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=2).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))
                ws.cell(row=cont, column=3).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))   
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))
                ws.row_dimensions[cont].height = 60
                #ponemos el otro recuadro
                cont +=2
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= True,color ="EAEAEA")
                ws.cell(row=cont, column=1).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=1).value = "AUTORIZA TRASLADO DE BIENES A BODEGA DE SANCTORUM"
                ws.row_dimensions[cont].height = 25
                #PONEMOS EL TEXTO NOMBRE 
                cont +=1
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"))
                ws.cell(row=cont, column=1).value = "NOMBRE:"
                ws.merge_cells(start_row=cont,start_column = 2, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"))
                ws.cell(row=cont, column=2).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=2).font = Font(name='Arial',size = 10, bold= True)
                ws.cell(row=cont, column=2).value = "LIC. " + autoridad

                ws.row_dimensions[cont].height = 25
                # PONEMOS EL TEXTO DE CARGO
                cont +=1
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"))
                ws.cell(row=cont, column=1).value = "CARGO:"
                ws.merge_cells(start_row=cont,start_column = 2, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"))
                ws.cell(row=cont, column=2).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=2).font = Font(name='Arial',size = 10, bold= True)
                ws.cell(row=cont, column=2).value = "DIRECTORA DE BIENES ASEGURADOS"    
                ws.row_dimensions[cont].height = 25
                #PONEMOS EL TEXTO DE FIRMA
                cont +=1
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"),bottom = Side(border_style='thin'))
                ws.cell(row=cont, column=1).value = "FIRMA:"
                ws.merge_cells(start_row=cont,start_column = 2, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=2).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))
                ws.cell(row=cont, column=3).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))   
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))
                ws.row_dimensions[cont].height = 60
                # CUADRO PARA QUIEN TRASLADA A LA BOVEDA
                cont +=2
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= True,color ="EAEAEA")
                ws.cell(row=cont, column=1).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=1).value = "TRASLADA A BODEGA DE BIENES ASEGURADOS"
                ws.row_dimensions[cont].height = 25
                #PONEMOS EL TEXTO NOMBRE 
                cont +=1
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"))
                ws.cell(row=cont, column=1).value = "NOMBRE:"
                ws.merge_cells(start_row=cont,start_column = 2, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"))
                ws.row_dimensions[cont].height = 25
                # PONEMOS EL TEXTO DE CARGO
                cont +=1
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"))
                ws.cell(row=cont, column=1).value = "CARGO:"
                ws.merge_cells(start_row=cont,start_column = 2, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"))    
                ws.row_dimensions[cont].height = 25
                #PONEMOS EL TEXTO DE FIRMA
                cont +=1
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"),bottom = Side(border_style='thin'))
                ws.cell(row=cont, column=1).value = "FIRMA:"
                ws.merge_cells(start_row=cont,start_column = 2, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=2).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))
                ws.cell(row=cont, column=3).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))   
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))
                ws.row_dimensions[cont].height = 60

                # FIN DEL CUADRO DE QUIEN TRASLADA

                # CUADRO DE QUIEN RECIBE EN LA BOVEDA DE BIENES ASEGURADOS
                cont +=2
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= True,color ="EAEAEA")
                ws.cell(row=cont, column=1).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=1).value = "RECIBE EN LA BODEGA DE BIENES ASEGURADOS"
                ws.row_dimensions[cont].height = 25
                #PONEMOS EL TEXTO NOMBRE 
                cont +=1
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"))
                ws.cell(row=cont, column=1).value = "NOMBRE:"
                ws.merge_cells(start_row=cont,start_column = 2, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"))
                ws.row_dimensions[cont].height = 25
                # PONEMOS EL TEXTO DE CARGO
                cont +=1
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"))
                ws.cell(row=cont, column=1).value = "CARGO:"
                ws.merge_cells(start_row=cont,start_column = 2, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"))    
                ws.row_dimensions[cont].height = 25
                #PONEMOS EL TEXTO DE FIRMA
                cont +=1
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"),bottom = Side(border_style='thin'))
                ws.cell(row=cont, column=1).value = "FIRMA:"
                ws.merge_cells(start_row=cont,start_column = 2, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=2).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))
                ws.cell(row=cont, column=3).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))   
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))
                ws.row_dimensions[cont].height = 60
                # FIN DEL CUADRO DE RECIBE EN BODEGA DE BIENES ASEGURADOS
                # PONEMOS EL CUADRO DE OBSERVACIONES
                cont += 2
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= True,color ="EAEAEA")
                ws.cell(row=cont, column=1).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=1).value = "OBSERVACIONES DEL INDICIO"
                ws.row_dimensions[cont].height = 25

                cont +=1
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))
                ws.cell(row=cont, column=2).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))
                ws.cell(row=cont, column=3).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))   
                ws.cell(row=cont, column=4).border = Border(right = Side(border_style="thin"),bottom = Side(border_style='thin'))
                ws.row_dimensions[cont].height = 60

                #configuramos la pagina
                ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
                ws.page_setup.fitToWidth = 60


                #Establecemos el nombre del archivo
                nombre_archivo ="ReporteTrasladoExcel.xlsx"
                #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
                response = HttpResponse(content_type="application/ms-excel")
                contenido = "attachment; filename={0}".format(nombre_archivo)
                response["Content-Disposition"] = contenido
                wb.save(response)
                return response

# clasee para enviar  e excel los ingresos por carpeta

class ReporteIndicios(TemplateView):

        #Usamos el método get para generar el archivo excel
        def get(self, request, *args, **kwargs):
                #Obtenemos todas las personas de nuestra base de datos

                campo = int(request.GET.get('campo'))
                indicios = Indicio.objects.filter(Ingreso_Num=campo)
                #Creamos el libro de trabajo
                wb = Workbook()
                #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
                ws = wb.active
                # pongo la imagen
                img = Image('media/imagen/fiscalia.png')
                ws.add_image(img, 'A1')
                # Hacer encabezado
                #le pongo el nombre de la hoja con la carpeta
                #ws.title = indicios.Ingreso_Num.Carpeta_Investigacion
                # para obtener el numero de carpeta, hora, fecha y quien entrega por una vez

                cuantos = 1
                for medio in indicios:
                        fecha= medio.Ingreso_Num.Fecha_Ingreso
                        hora = medio.Ingreso_Num.Hora_Ingreso
                        carpeta = medio.Ingreso_Num.Carpeta_Investigacion
                        persona = medio.Persona_Entrega
                        oficio = medio.Ingreso_Num.Oficio
                        cadena = medio.Cadena
                        embalaje = medio.Embalaje
                        observa = medio.Observaciones
                        cuantos = cuantos + 1

                # saber el embalaje y que tipo de cadena tiene el indicio
                if cadena == 'S':
                    cadena = 'CON CADENA DE CUSTODIA'
                elif cadena == 'N':
                    cadena = 'SIN CADENA DE CUSTODIA'
                elif cadena == 'CS':
                    cadena = 'COPIA SIMPLE'
                else:
                    cadena = 'COPIA CERTIFICADA'

                """
                B', 'Bolsa de Plástico'),
                     ('BP', 'Bolsa de Papel'),
                     ('SB', 'Sobre Blanco'),
                     ('SM', 'Sobre Manila'),
                     ('EM', 'Emplayado'),
                     ('CJ', 'Caja de Cartón'),
                     ('SE', 'Sin Embalaje'),
                     ('O', 'Otro'
                """
                if embalaje == 'B':
                    embalaje = 'BOLSA DE PLASTICO'
                elif embalaje == 'BP':
                    embalaje = 'BOLSA DE PAPEL'
                elif embalaje == 'SB':
                    embalaje = 'SOBRE BLANCO'
                elif embalajen == 'SM':
                    embalaje = 'SOBRE MANILA'
                elif embalaje == 'EM':
                    embalaje = 'EMPLAYADO'
                elif embalaje == 'CJ':
                    embalaje = 'CAJA DE CARTON'
                elif embalaje == 'SE':
                    embalaje = 'SIN EMBALAJE'
                else:
                    embalaje = 'OTRO'



                # nombre de la hoja como la carpeta
                #date_object = datetime.strptime(fecha, '%dd/%mm/%Y')
                fecha = datetime.strftime(fecha, '%d/%m/%Y')       
                ws.title = 'indicios'
                #En la celda C2 ponemos el texto 'FISCALIA GENERAL DEL ESTADO' y la decoramos
                #alignment=Alignment
                ws['C3'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C3'].font = Font(name='Arial',size = 14, bold= True)
                ws['C3'] = 'Fiscalía General del Estado'
                #En la celda C3 ponemos el texto 'Oficialía Mayor' y la decoramos
                ws['C4'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C4'].font = Font(name='Arial',size = 11, bold= True, color = "BFBDA4")
                ws['C4'] = 'Oficialía Mayor'
                # en la celda c4 ponemos el texto "Dirección de Bienes Asegurados"
                ws['C5'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C5'].font = Font(name='Arial',size = 11, bold= True)
                ws['C5'] = 'Dirección de Bienes Asegurados'
                # en la celda c8 ponemos el texto "Dirección de Bienes Asegurados"
                ws['C7'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C7'].font = Font(name='Arial',size = 12, bold= True)
                ws['C7'] = 'INGRESO'
                # En la celda D6 pones el titulo "Carpeta de Investigacion" con relleno y bordes
                ws['D6'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['D6'].font = Font(name='Arial',size = 10, bold= True,color ="EAEAEA")
                ws['D6'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['D6'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['D6'] = 'CARPETA DE INVESTIGACION'
                # En la celda D7 pones la variable carpeta sin  relleno y bordes
                ws['D7'].alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws['D7'].font = Font(name='Arial',size = 10, bold= False)
                ws['D7'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=7,column=4).value = carpeta
                # hacemos en la celda a9 el encabezado "folio"
                ws.merge_cells('A9:B9')
                ws['A9'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['A9'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['A9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),
                                        top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['A9'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['A9'] = 'FOLIO'
                # En la celda A10 pones la variable campo que es el id de la carpeta sin  relleno y bordes
                ws.merge_cells('A10:B10')
                ws['A10'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['A10'].font = Font(name='Arial',size = 10, bold= False)
                ws['A10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['B10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=10,column=1).value = campo
                # hacemos en la celda c9 el encabezado "lugar de entrega recepcion"
                ws['C9'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C9'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['C9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['C9'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['C9'] = 'LUGAR DE LA ENTREGA - RECEPCIÓN'
                # En la celda A10 pones la variable campo que es el id de la carpeta sin  relleno y bordes
                ws['C10'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C10'].font = Font(name='Arial',size = 10, bold= False)
                ws['C10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['C10'] = 'BOVEDA DE INDICIOS'
                # hacemos en la celda D9 el encabezado "Fecha"
                ws['D9'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['D9'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['D9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['D9'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['D9'] = 'FECHA'
                # En la celda D10 pones la variable campo que es el id de la carpeta sin  relleno y bordes
                ws['D10'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['D10'].font = Font(name='Arial',size = 10, bold= False)
                ws['D10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=10,column=4).value = fecha
                # hacemos en la celda E9 el encabezado "Hora"
                ws['E9'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['E9'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['E9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['E9'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['E9'] = 'HORA'
                # En la celda D10 pones la variable campo que es el id de la carpeta sin  relleno y bordes
                ws['E10'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['E10'].font = Font(name='Arial',size = 10, bold= False)
                ws['E10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=10,column=5).value = hora 
                # hacemos en la celda F9 el encabezado "OFICO"
                ws['F9'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['F9'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['F9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['F9'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['F9'] = 'OFICIO'
                # En la celda D10 pones la variable campo que es el id de la carpeta sin  relleno y bordes
                ws['F10'].alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws['F10'].font = Font(name='Arial',size = 10, bold= False)
                ws['F10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=10,column=6).value = oficio
                # pomer texto en la celda a12
                ws.merge_cells('A12:D12')
                ws['A12'].alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws['A12'].font = Font(name='Arial',size = 10, bold= False)
                ws['A12'] = '1. Inventario (escribe el número, letra o combinación alfanumérica con la que se identifica a cada indicio o elemento material probatorio que se entrega, así como su tipo o clase. Cancele los espacios sobrantes'
                # HACER EL ENCABEZADO
                #ID
                ws['A14'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['A14'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['A14'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['A14'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['A14'] = 'ID'
                #TIPO
                ws['B14'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['B14'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['B14'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['B14'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['B14'] = 'TIPO'
                #DESCRIPCION
                ws['C14'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C14'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['C14'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['C14'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['C14'] = 'DESCRIPCION'
                #RUE
                ws['D14'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['D14'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['D14'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['D14'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['D14'] = 'RUE'
                # definir la anchura para las columnas a,b,c,d,e,y f
                ws.column_dimensions['A'].width = 7
                ws.column_dimensions['B'].width = 7
                ws.column_dimensions['C'].width = 76
                ws.column_dimensions['D'].width = 30
                ws.column_dimensions['E'].width = 11
                
                # definir la altura para las filas
                
                ws.row_dimensions[7].height = 35
                ws.row_dimensions[10].height = 35
                ws.row_dimensions[12].height = 50
                ws.row_dimensions[14].height = 26
                cont=15
                
                #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
                for indicio in indicios:
                        ws.cell(row=cont,column=1).value = indicio.No_Indicio
                        ws.cell(row=cont,column=2).value = indicio.Tipo
                        ws.cell(row=cont,column=3).value = indicio.Descripcion
                        ws.cell(row=cont,column=4).value = indicio.Rue
                        #ponemos el diseño de la celda, el tipo de letra  y los borde
                        ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                        ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                        ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                        # columna dos
                        ws.cell(row=cont, column=2).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                        ws.cell(row=cont, column=2).font = Font(name='Arial',size = 10, bold= False)
                        ws.cell(row=cont, column=2).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                        # columna 3
                        ws.cell(row=cont, column=3).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                        ws.cell(row=cont, column=3).font = Font(name='Arial',size = 10, bold= False)
                        ws.cell(row=cont, column=3).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                        # columna 4
                        ws.cell(row=cont, column=4).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                        ws.cell(row=cont, column=4).font = Font(name='Arial',size = 10, bold= False)
                        ws.cell(row=cont, column=4).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                        ws.row_dimensions[cont].height = 60
                        cont += 1
                        
                
                # ponemos el concepto de cadena
                
                ws.row_dimensions[cont].height = 25
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))        

                ws.cell(row=cont, column=2).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=2).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=2).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                
                ws.cell(row=cont, column=3).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=3).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=3).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))

                ws.cell(row=cont, column=4).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=4).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=4).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont,column=3).value = cadena
                # el mensaje del tipo de embalaje

                cont +=1

                
                
                ws.row_dimensions[cont].height = 45
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= True)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont,column=1).value = "2. Embalaje (Señale las condiciones en las que se encuentran los embales. Cuando alguno de ellos presente alteración, deterioro o cualquier otra anomalía, especifique dicha condición)."
                ws.cell(row=cont, column=2).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=3).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=4).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                # ponemos el tipo de embalaje
                cont += 1
                ws.row_dimensions[cont].height = 45
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= True)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont,column=1).value = embalaje + '  Observaciones: '+ observa
                ws.cell(row=cont, column=2).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=3).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=4).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                #PONEMOS LAS FIRMAS DE LOS TITULARES
                cont += 2
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 11, bold= True)
                ws.cell(row=cont,column=1).value = "ENTREGA"
                cont += 3
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 11, bold= True)
                ws.cell(row=cont,column=1).value = persona
                cont += 1
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 11, bold= False)
                ws.cell(row=cont,column=1).value = "ADSCRITO A: "
                cont += 3
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 11, bold= True)
                ws.cell(row=cont,column=1).value = "RECIBE"
                cont += 3
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 11, bold= True)
                ws.cell(row=cont,column=1).value = "NOMBRE DE QUIEN RECIBE:"
                cont += 1
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 11, bold= False)
                ws.cell(row=cont,column=1).value = "ANALISTA DE LA DIRECCION DE BIENES ASEGURADOS "
                cont += 2
                ws.cell(row=cont, column = 1).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws.cell(row=cont, column = 2).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws.cell(row=cont, column = 3).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws.cell(row=cont, column = 4).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                cont += 1
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont,column=1).value = "ESTE FORMATO CONSTITUYE UN RECIBO PERSONAL, POR LO QUE SE LE SUGIERE CONSERVAR UNA COPIA."

                #configuramos la pagina
                ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
                ws.page_setup.fitToWidth = 60


                #Establecemos el nombre del archivo
                nombre_archivo ="ReporteIndiciosExcel.xlsx"
                #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
                response = HttpResponse(content_type="application/ms-excel")
                contenido = "attachment; filename={0}".format(nombre_archivo)
                response["Content-Disposition"] = contenido
                wb.save(response)
                return response

# fin de la clase

# CLASE PARA EL REPORTE A EXCEL DE LAS SALIDAS DEFINITIVAS
class ReporteDefinitiva(TemplateView):

        #Usamos el método get para generar el archivo excel
        def get(self, request, *args, **kwargs):
                #Obtenemos todas las personas de nuestra base de datos

                campo = int(request.GET.get('campo'))
                indicios = Definitiva.objects.filter(Carpeta=campo)
                #Creamos el libro de trabajo
                wb = Workbook()
                #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
                ws = wb.active
                # pongo la imagen
                img = Image('media/imagen/fiscalia.png')
                ws.add_image(img, 'A1')
                # Hacer encabezado
                #le pongo el nombre de la hoja con la carpeta
                #ws.title = indicios.Ingreso_Num.Carpeta_Investigacion
                # para obtener el numero de carpeta, hora, fecha y quien entrega por una vez

                cuantos = 1
                for medio in indicios:
                        fecha= medio.Fecha_de_Movimiento
                        hora = medio.Carpeta.Hora_Ingreso
                        carpeta = medio.Carpeta.Carpeta_Investigacion
                        persona = medio.Persona_Recibe
                        oficio = medio.Oficio
                        cadena = medio.Rue.Cadena
                        embalaje = medio.Rue.Embalaje
                        observa = medio.Rue.Observaciones
                        cuantos = cuantos + 1

                # saber el embalaje y que tipo de cadena tiene el indicio
                if cadena == 'S':
                    cadena = 'CON CADENA DE CUSTODIA'
                elif cadena == 'N':
                    cadena = 'SIN CADENA DE CUSTODIA'
                elif cadena == 'CS':
                    cadena = 'COPIA SIMPLE'
                else:
                    cadena = 'COPIA CERTIFICADA'

                """
                B', 'Bolsa de Plástico'),
                     ('BP', 'Bolsa de Papel'),
                     ('SB', 'Sobre Blanco'),
                     ('SM', 'Sobre Manila'),
                     ('EM', 'Emplayado'),
                     ('CJ', 'Caja de Cartón'),
                     ('SE', 'Sin Embalaje'),
                     ('O', 'Otro'
                """
                if embalaje == 'B':
                    embalaje = 'BOLSA DE PLASTICO'
                elif embalaje == 'BP':
                    embalaje = 'BOLSA DE PAPEL'
                elif embalaje == 'SB':
                    embalaje = 'SOBRE BLANCO'
                elif embalajen == 'SM':
                    embalaje = 'SOBRE MANILA'
                elif embalaje == 'EM':
                    embalaje = 'EMPLAYADO'
                elif embalaje == 'CJ':
                    embalaje = 'CAJA DE CARTON'
                elif embalaje == 'SE':
                    embalaje = 'SIN EMBALAJE'
                else:
                    embalaje = 'OTRO'



                # nombre de la hoja como la carpeta
                #date_object = datetime.strptime(fecha, '%dd/%mm/%Y')
                fecha = datetime.strftime(fecha, '%d/%m/%Y')       
                ws.title = 'salida definitiva'
                #En la celda C2 ponemos el texto 'FISCALIA GENERAL DEL ESTADO' y la decoramos
                #alignment=Alignment
                ws['C3'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C3'].font = Font(name='Arial',size = 14, bold= True)
                ws['C3'] = 'Fiscalía General del Estado'
                #En la celda C3 ponemos el texto 'Oficialía Mayor' y la decoramos
                ws['C4'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C4'].font = Font(name='Arial',size = 11, bold= True, color = "BFBDA4")
                ws['C4'] = 'Oficialía Mayor'
                # en la celda c4 ponemos el texto "Dirección de Bienes Asegurados"
                ws['C5'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C5'].font = Font(name='Arial',size = 11, bold= True)
                ws['C5'] = 'Dirección de Bienes Asegurados'
                # en la celda c8 ponemos el texto "Dirección de Bienes Asegurados"
                ws['C7'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C7'].font = Font(name='Arial',size = 12, bold= True)
                ws['C7'] = 'DEVOLUCION DEFINITIVA'
                # En la celda D6 pones el titulo "Carpeta de Investigacion" con relleno y bordes
                ws['D6'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['D6'].font = Font(name='Arial',size = 10, bold= True,color ="EAEAEA")
                ws['D6'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['D6'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['D6'] = 'CARPETA DE INVESTIGACION'
                # En la celda D7 pones la variable carpeta sin  relleno y bordes
                ws['D7'].alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws['D7'].font = Font(name='Arial',size = 10, bold= False)
                ws['D7'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=7,column=4).value = carpeta
                # hacemos en la celda a9 el encabezado "folio"
                ws.merge_cells('A9:B9')
                ws['A9'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['A9'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['A9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),
                                        top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['A9'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['A9'] = 'FOLIO'
                # En la celda A10 pones la variable campo que es el id de la carpeta sin  relleno y bordes
                ws.merge_cells('A10:B10')
                ws['A10'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['A10'].font = Font(name='Arial',size = 10, bold= False)
                ws['A10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['B10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=10,column=1).value = campo
                # hacemos en la celda c9 el encabezado "lugar de entrega recepcion"
                ws['C9'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C9'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['C9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['C9'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['C9'] = 'LUGAR DE LA ENTREGA - RECEPCIÓN'
                # En la celda A10 pones la variable campo que es el id de la carpeta sin  relleno y bordes
                ws['C10'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C10'].font = Font(name='Arial',size = 10, bold= False)
                ws['C10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['C10'] = 'BOVEDA DE INDICIOS'
                # hacemos en la celda D9 el encabezado "Fecha"
                ws['D9'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['D9'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['D9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['D9'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['D9'] = 'FECHA'
                # En la celda D10 pones la variable campo que es el id de la carpeta sin  relleno y bordes
                ws['D10'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['D10'].font = Font(name='Arial',size = 10, bold= False)
                ws['D10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=10,column=4).value = fecha
                # hacemos en la celda E9 el encabezado "Hora"
                ws['E9'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['E9'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['E9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['E9'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['E9'] = 'HORA'
                # En la celda D10 pones la variable campo que es el id de la carpeta sin  relleno y bordes
                ws['E10'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['E10'].font = Font(name='Arial',size = 10, bold= False)
                ws['E10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=10,column=5).value = hora 
                # hacemos en la celda F9 el encabezado "OFICO"
                ws['F9'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['F9'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['F9'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['F9'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['F9'] = 'OFICIO'
                # En la celda D10 pones la variable campo que es el id de la carpeta sin  relleno y bordes
                ws['F10'].alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws['F10'].font = Font(name='Arial',size = 10, bold= False)
                ws['F10'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=10,column=6).value = oficio
                # pomer texto en la celda a12
                ws.merge_cells('A12:D12')
                ws['A12'].alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws['A12'].font = Font(name='Arial',size = 10, bold= False)
                ws['A12'] = '1. Inventario (escribe el número, letra o combinación alfanumérica con la que se identifica a cada indicio o elemento material probatorio que se entrega, así como su tipo o clase. Cancele los espacios sobrantes'
                # HACER EL ENCABEZADO
                #ID
                ws['A14'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['A14'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['A14'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['A14'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['A14'] = 'ID'
                #TIPO
                ws['B14'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['B14'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['B14'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['B14'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['B14'] = 'TIPO'
                #DESCRIPCION
                ws['C14'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['C14'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['C14'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['C14'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['C14'] = 'DESCRIPCION'
                #RUE
                ws['D14'].alignment = Alignment(horizontal = "center",vertical= "center")
                ws['D14'].font = Font(name='Arial',size = 11, bold= True,color ="EAEAEA")
                ws['D14'].border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws['D14'].fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws['D14'] = 'RUE'
                # definir la anchura para las columnas a,b,c,d,e,y f
                ws.column_dimensions['A'].width = 7
                ws.column_dimensions['B'].width = 7
                ws.column_dimensions['C'].width = 76
                ws.column_dimensions['D'].width = 30
                ws.column_dimensions['E'].width = 11
                
                # definir la altura para las filas
                
                ws.row_dimensions[7].height = 35
                ws.row_dimensions[10].height = 35
                ws.row_dimensions[12].height = 50
                ws.row_dimensions[14].height = 26
                cont=15
                
                #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
                for indicio in indicios:
                        ws.cell(row=cont,column=1).value = indicio.Rue.No_Indicio
                        ws.cell(row=cont,column=2).value = indicio.Rue.Tipo
                        ws.cell(row=cont,column=3).value = indicio.Rue.Descripcion
                        ws.cell(row=cont,column=4).value = indicio.Rue.Rue
                        #ponemos el diseño de la celda, el tipo de letra  y los borde
                        ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                        ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                        ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                        # columna dos
                        ws.cell(row=cont, column=2).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                        ws.cell(row=cont, column=2).font = Font(name='Arial',size = 10, bold= False)
                        ws.cell(row=cont, column=2).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                        # columna 3
                        ws.cell(row=cont, column=3).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                        ws.cell(row=cont, column=3).font = Font(name='Arial',size = 10, bold= False)
                        ws.cell(row=cont, column=3).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                        # columna 4
                        ws.cell(row=cont, column=4).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                        ws.cell(row=cont, column=4).font = Font(name='Arial',size = 10, bold= False)
                        ws.cell(row=cont, column=4).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                        ws.row_dimensions[cont].height = 60
                        cont += 1
                        
                
                # ponemos el concepto de cadena
                
                ws.row_dimensions[cont].height = 25
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))        

                ws.cell(row=cont, column=2).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=2).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=2).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                
                ws.cell(row=cont, column=3).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=3).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=3).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))

                ws.cell(row=cont, column=4).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=4).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont, column=4).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont,column=3).value = cadena
                # el mensaje del tipo de embalaje

                cont +=1

                
                
                ws.row_dimensions[cont].height = 45
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= True)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont,column=1).value = "2. Embalaje (Señale las condiciones en las que se encuentran los embales. Cuando alguno de ellos presente alteración, deterioro o cualquier otra anomalía, especifique dicha condición)."
                ws.cell(row=cont, column=2).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=3).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=4).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                # ponemos el tipo de embalaje
                cont += 1
                ws.row_dimensions[cont].height = 45
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= True)
                ws.cell(row=cont, column=1).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont,column=1).value = embalaje + '  Observaciones: '+ observa
                ws.cell(row=cont, column=2).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=3).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                ws.cell(row=cont, column=4).border = Border(left = Side(border_style= "thin"), right = Side(border_style="thin" ),top= Side(border_style="thin"), bottom = Side(border_style="thin"))
                #PONEMOS LAS FIRMAS DE LOS TITULARES
                cont += 2
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 11, bold= True)
                ws.cell(row=cont,column=1).value = "ENTREGA"
                cont += 3
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 11, bold= True)
                ws.cell(row=cont,column=1).value = "NOMBRE DE QUIEN ENTREGA"
                cont += 1
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 11, bold= False)
                ws.cell(row=cont,column=1).value = "ANALISTA DE LA DIRECCION DE BIENES ASEGURADOS "
                cont += 3
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 11, bold= True)
                ws.cell(row=cont,column=1).value = "RECIBE"
                cont += 3
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 11, bold= True)
                ws.cell(row=cont,column=1).value = persona
                cont += 1
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 11, bold= False)
                ws.cell(row=cont,column=1).value = "CARGO  "
                cont += 3
                ws.cell(row=cont, column = 1).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws.cell(row=cont, column = 2).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws.cell(row=cont, column = 3).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                ws.cell(row=cont, column = 4).fill = PatternFill(start_color = '00305A', end_color = '00305A', fill_type = 'solid')
                cont += 1
                ws.merge_cells(start_row=cont,start_column = 1, end_row = cont, end_column = 4)
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal = "center",vertical= "center",wrap_text=True)
                ws.cell(row=cont, column=1).font = Font(name='Arial',size = 10, bold= False)
                ws.cell(row=cont,column=1).value = "ESTE FORMATO CONSTITUYE UN RECIBO PERSONAL, POR LO QUE SE LE SUGIERE CONSERVAR UNA COPIA."

                #configuramos la pagina
                ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
                ws.page_setup.fitToWidth = 60


                #Establecemos el nombre del archivo
                nombre_archivo ="ReporteDefinitivoExcel.xlsx"
                #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
                response = HttpResponse(content_type="application/ms-excel")
                contenido = "attachment; filename={0}".format(nombre_archivo)
                response["Content-Disposition"] = contenido
                wb.save(response)
                return response

# FIN DE LA CLASEE DE SALIDAS DEFINITIVAS




def indicioscarpeta(request):
       todosindicios = Indicio.objects.all()
       carpeta = None
       if request.method == 'POST':
        form = SelExpedienteForm(data=request.POST)
        print(form)
        print("si entro al if del POST")
        #print(form.data['Ingreso_Num']) 

        if form.is_valid():
            print('si entro al if del is_valid')
            indicio_id = form.data['Ingreso_Num']
            # aqui lo busco en la carpeta de expedientes y paso los datos
            carpeta = Expediente.objects.get(pk=indicio_id)
            todosindicios = Indicio.objects.filter(Ingreso_Num=indicio_id)
            print('no entro al is_valid')
        else:
            error=form['Ingreso_Num'].errors
            print('no entro AL is_valid' + error)

            # aqui obtengo el id del expediente
            
        """if 'excel' in request.POST:
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=Report.xlsx'
            xlsx_data = WriteToExcel(weather_period, town)
            response.write(xlsx_data)
            return response
            if 'pdf' in request.POST:
            response = HttpResponse(content_type='application/pdf')
            today = date.today()
            filename = 'pdf_demo' + today.strftime('%Y-%m-%d')
            response['Content-Disposition'] =\
                'attachement; filename={0}.pdf'.format(filename)
            buffer = BytesIO()
            report = PdfPrint(buffer, 'A4')
            pdf = report.report(weather_period, 'Weather statistics data')
            response.write(pdf)
            return response
            """
       else:
        print('no entro al if del POST')
        form = SelExpedienteForm()
        print(form)
        print('imprimio el form sin entar al if')

       template_name = "boveda/indicios_list1.html"
       context = {
        'form': form,
        'carpeta': carpeta,
        'todosindicios': todosindicios,
       }
       return render(request,template_name,context)

def buscacarpeta(request):
    return render(request,'boveda/busca_carpeta.html')

def buscar(request):
    error = False
    if 'q' in request.GET:
        q = request.GET['q']
        if not q:
            error = True
        else:
            indicios = Indicio.objects.filter(Rue__icontains=q)
            #print(indicios)
            return render(request, 'boveda/resultado_carpeta.html', {'indicios': indicios,'query':q})
    return render(request, 'boveda/busca_carpeta.html',{'error':error})

